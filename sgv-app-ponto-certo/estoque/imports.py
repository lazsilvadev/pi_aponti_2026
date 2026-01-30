"""Helpers para processar importação de arquivos (CSV / Excel).

Funções: parse_file_to_records, process_import

Este módulo é intencionalmente livre de efeitos colaterais sobre `produtos`.
Ele apenas lê/normaliza/valida e retorna listas de itens a serem inseridos
e a lista de nomes/códigos duplicados detectados.
"""

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _norm(s: str) -> str:
    import unicodedata as _ud

    if not s:
        return ""
    nf = _ud.normalize("NFD", str(s))
    return "".join(c for c in nf if _ud.category(c) != "Mn").strip().lower()


def _norm_cb(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s.replace(".0", "").isdigit():
        s = s[:-2]
    return s


def parse_file_to_records(
    file_path: str,
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
    p = Path(file_path)
    ext = p.suffix.lower()
    if ext in (".xls", ".xlsx"):
        # tenta usar pandas (lazy import) e fallback para leitura manual/openpyxl
        try:
            import pandas as pd

            df = pd.read_excel(file_path, engine="openpyxl")
        except Exception as ex_openpyxl:
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                headers = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in next(ws.iter_rows(max_row=1))
                ]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    d = {}
                    for i, val in enumerate(row):
                        key = headers[i] if i < len(headers) else f"col{i}"
                        d[key] = val
                    rows.append(d)
                try:
                    import pandas as pd

                    df = pd.DataFrame(rows)
                except Exception:
                    # criar records manualmente
                    records = []
                    for r in rows:
                        rec = {k: ("" if v is None else v) for k, v in r.items()}
                        records.append(rec)
                    return records, None
            except Exception:
                try:
                    import pandas as pd

                    df = pd.read_excel(file_path, engine="calamine")
                except Exception as ex_calamine:
                    return None, (
                        f"Falha ao ler Excel: openpyxl({ex_openpyxl}), calamine({ex_calamine})."
                    )

        # normaliza DataFrame para lista de dicionários mantendo números
        records: List[Dict[str, Any]] = []
        import pandas as _pd

        for row in df.to_dict(orient="records"):
            rec: Dict[str, Any] = {}
            for k, v in row.items():
                key = str(k).strip()
                if v is None or (isinstance(v, float) and _pd.isna(v)):
                    rec[key] = ""
                elif hasattr(v, "strftime"):
                    try:
                        rec[key] = v.strftime("%d/%m/%Y")
                    except Exception:
                        rec[key] = str(v)
                elif isinstance(v, (int, float)):
                    rec[key] = v
                else:
                    rec[key] = str(v)
            records.append(rec)
        return records, None
    else:
        try:
            with open(file_path, encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                return list(reader), None
        except Exception as e:
            return None, str(e)


def process_import(
    file_path: str,
    produtos: List[Dict[str, Any]],
    converter_texto_para_data,
    converter_texto_para_preco,
    validar_produto,
) -> Tuple[List[Dict[str, Any]], List[str], Optional[str]]:
    """Processa arquivo de importação e retorna (itens_para_inserir, duplicados, erro)."""
    records, err = parse_file_to_records(file_path)
    if err:
        return [], [], f"Erro ao ler arquivo: {err}"
    if records is None:
        return [], [], "Arquivo vazio ou inválido"

    existentes_cod = {
        _norm_cb(p.get("codigo_barras")) for p in produtos if p.get("codigo_barras")
    }
    existentes_nome_cat = {
        (_norm(p.get("nome")), _norm(p.get("categoria"))) for p in produtos
    }

    importados_dedup: List[Dict[str, Any]] = []
    duplicados: List[str] = []

    for row in records:
        try:
            nome = row.get("Nome") or row.get("nome")
            categoria = row.get("Categoria") or row.get("categoria")
            validade = row.get("Validade") or row.get("validade")
            quantidade = row.get("Quantidade") or row.get("quantidade")
            codigo_barras = (
                row.get("Código de Barras")
                or row.get("codigo_barras")
                or row.get("Codigo de Barras")
            )
            validade_obj = converter_texto_para_data(validade)
            qtd = validar_produto(nome, categoria, quantidade, validade_obj)
            preco_venda = converter_texto_para_preco(
                row.get("Preço")
                or row.get("Preco")
                or row.get("preco_venda")
                or row.get("preco")
                or row.get("Preço de Venda")
            )
            preco_custo = converter_texto_para_preco(
                row.get("Preço de Custo")
                or row.get("Preco de Custo")
                or row.get("Custo")
                or row.get("custo")
                or row.get("preco_custo")
            )
            cb = _norm_cb(codigo_barras)
            key_nome_cat = (_norm(nome), _norm(categoria))

            if (
                (cb and cb in existentes_cod)
                or (key_nome_cat in existentes_nome_cat)
                or (
                    (
                        cb
                        and any(
                            str(n.get("codigo_barras")) == cb for n in importados_dedup
                        )
                    )
                    or (
                        key_nome_cat
                        in {
                            (_norm(n.get("nome")), _norm(n.get("categoria")))
                            for n in importados_dedup
                        }
                    )
                )
            ):
                duplicados.append(nome or cb or "(sem nome)")
                continue

            importados_dedup.append(
                {
                    "nome": nome,
                    "categoria": categoria,
                    "validade": validade_obj,
                    "quantidade": qtd,
                    "preco_venda": preco_venda,
                    "preco_custo": preco_custo,
                    "codigo_barras": cb,
                }
            )
        except Exception:
            # ignora linha inválida
            continue

    return importados_dedup, duplicados, None
